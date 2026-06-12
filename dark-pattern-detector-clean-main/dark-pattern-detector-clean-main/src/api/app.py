from flask import Flask, jsonify, request, send_file

from flask_cors import CORS
from joblib import load
import pandas as pd
import base64
import openpyxl
from openpyxl.drawing.image import Image
from io import BytesIO
from zipfile import ZipFile
import json
import numpy as np

presence_classifier = load('presence_classifer.joblib')
presence_vect = load('presence_vectorizer.joblib')
category_classifier = load('category_classifier.joblib')
category_vect = load('category_vectorizer.joblib')

app = Flask(__name__)
CORS(app)


@app.route('/', methods=['POST', 'GET'])
def main():
    if request.method == 'POST':
        try:
            output = []
            data = request.get_json().get('tokens')

            for token in data:
                result = presence_classifier.predict(
                    presence_vect.transform([token]))[0]
                if result == 'Dark':
                    cat = category_classifier.predict(
                        category_vect.transform([token]))[0]
                    # Convert numpy string to native Python string using item()
                    if hasattr(cat, 'item'):
                        output.append(cat.item())
                    else:
                        output.append(str(cat))
                else:
                    # Convert numpy string to native Python string using item()
                    if hasattr(result, 'item'):
                        output.append(result.item())
                    else:
                        output.append(str(result))

            dark = [data[i] for i in range(len(output)) if output[i] == 'Dark']
            for d in dark:
                print(d)
            print()
            print(len(dark))

            # Convert all items to native Python strings explicitly
            result_list = [str(item) if not isinstance(item, str) else item for item in output]
            
            # Ensure all items are native Python strings using json.dumps then loads
            json_str = json.dumps({'result': result_list})
            return json_str, 200, {'Content-Type': 'application/json'}
        except Exception as e:
            print(f"Error in main: {str(e)}")
            import traceback
            traceback.print_exc()
            return jsonify({'error': str(e)}), 500
    else:
        return jsonify({'message': 'POST request expected'}), 400


@app.route('/report', methods=['POST'])
def report():
    if request.method == 'POST':
        data = request.get_json()
        image_data = data.get('image')
        description = data.get('description')

        image_bytes = base64.b64decode(image_data.split(',')[1])

        try:
            workbook = openpyxl.load_workbook('reports.xlsx')
            csv_file_path = 'path/to/your/dataset.csv'
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        worksheet = workbook.active
        if worksheet.title != 'Reports':
            worksheet = workbook.create_sheet(title='Reports')

        
        worksheet.append([description])

        image_stream = BytesIO(image_bytes)

        img = Image(image_stream)
        worksheet.add_image(img, f'B{worksheet.max_row + 1}')

        workbook.save('reports.xlsx')

        return jsonify({'message': 'Report saved successfully.'})





if __name__ == '__main__':
    app.run(threaded=True, debug=True)
